"""
Excel builder — generates the Tally Audit Pack XLSX with all sheets.
Uses openpyxl for full formatting control.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Style constants ──────────────────────────────────────────────────────────
C_BG_DARK   = "0D0D0D"
C_BG_HEAD   = "1A1A2E"
C_ACCENT    = "F4A100"
C_ACCENT2   = "3B82F6"
C_GREEN     = "22C55E"
C_RED       = "EF4444"
C_WHITE     = "E8E8E8"
C_MUTED     = "888888"
C_SEC_BG    = "16213E"
C_TOTAL_BG  = "0F3460"
C_ALT_ROW   = "111827"

def hdr_font(size=10, bold=True, color=C_WHITE):
    return Font(name="Arial", size=size, bold=bold, color=color)

def body_font(size=9, bold=False, color=C_WHITE):
    return Font(name="Arial", size=size, bold=bold, color=color)

def mono_font(size=9, color=C_WHITE):
    return Font(name="Courier New", size=size, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="2D2D2D")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center")

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def right():
    return Alignment(horizontal="right", vertical="center")

NUM_FMT = '#,##0.00;[RED]-#,##0.00;"-"'

# ── Sheet helpers ────────────────────────────────────────────────────────────

def write_company_header(ws, company: str, title: str, subtitle: str = ""):
    """Write a 3-row company header block and return the next row index."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    c = ws.cell(1, 1, company)
    c.font = Font(name="Arial", size=14, bold=True, color=C_ACCENT)
    c.fill = fill(C_BG_HEAD)
    c.alignment = left()

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    c = ws.cell(2, 1, title)
    c.font = Font(name="Arial", size=11, bold=True, color=C_WHITE)
    c.fill = fill(C_BG_HEAD)
    c.alignment = left()

    if subtitle:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10)
        c = ws.cell(3, 1, subtitle)
        c.font = Font(name="Arial", size=9, bold=False, color=C_MUTED)
        c.fill = fill(C_BG_HEAD)
        c.alignment = left()

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 14
    return 5  # data starts at row 5


def write_col_headers(ws, row: int, headers: list, col_widths: list = None):
    """Write a column header row with dark background."""
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row, ci, h)
        c.font  = hdr_font(9, True, C_WHITE)
        c.fill  = fill(C_BG_HEAD)
        c.alignment = center() if ci > 1 else left()
        c.border = thin_border()
    ws.row_dimensions[row].height = 16

    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w


def write_data_row(ws, row: int, values: list, is_section=False, is_total=False, is_alt=False):
    bg = C_TOTAL_BG if is_total else (C_SEC_BG if is_section else (C_ALT_ROW if is_alt else "111111"))
    for ci, v in enumerate(values, 1):
        c = ws.cell(row, ci, v)
        c.fill = fill(bg)
        c.border = thin_border()
        c.alignment = right() if isinstance(v, (int, float)) and ci > 1 else left()
        if is_total:
            c.font = Font(name="Arial", size=9, bold=True, color=C_GREEN)
        elif is_section:
            c.font = Font(name="Arial", size=9, bold=True, color=C_ACCENT)
        else:
            c.font = body_font()
        if isinstance(v, (int, float)) and ci > 1 and v != 0:
            c.number_format = NUM_FMT
    ws.row_dimensions[row].height = 14


# ── Classification groups ────────────────────────────────────────────────────
PL_INCOME  = {"income", "other_income"}
PL_EXPENSE = {"purchase", "empl_expense", "indir_expense"}
BS_LIAB    = {"equity","reserve","lt_loan","provision","creditor","branch","current_liab","tax_gst"}
BS_ASSET   = {"fixed_asset","investment","deposit","debtor","bank","advance","current_asset","tax_tds","other","tax_it"}


def fiscal_month(date_str: str) -> int:
    """Returns 0=Apr … 11=Mar, or -1 if unparseable."""
    try:
        if not date_str:
            return -1
        # DD-MM-YYYY
        parts = str(date_str).replace("/", "-").split("-")
        if len(parts) == 3:
            m = int(parts[1])
            return m - 4 if m >= 4 else m + 8
    except Exception:
        pass
    return -1


# ── Main builder ─────────────────────────────────────────────────────────────

def build_excel(
    company, from_date, to_date, entity_type, currency,
    vouchers, ledgers, groups, trial_balance,
    edit_log, selected_sheets, log_fn
) -> bytes:

    wb = Workbook()
    wb.remove(wb.active)          # remove default sheet
    period = f"{from_date} to {to_date}"
    total_sheets = len(selected_sheets)
    done = 0

    def pct():
        return int(20 + (done / max(total_sheets, 1)) * 75)

    # ── Helper: filter & compute totals ──────────────────────────────────────
    def tb_total(cls_set, side="cr"):
        t = 0.0
        for v in trial_balance.values():
            if v["cls"] in cls_set:
                t += v.get("closing_cr" if side=="cr" else "closing_dr", 0) or 0
        return t

    # Enrich trial balance with closing balances
    for name, t in trial_balance.items():
        cl_raw = t.get("op_dr",0) - t.get("op_cr",0) + t.get("dr",0) - t.get("cr",0)
        t["closing_dr"] = cl_raw if cl_raw >= 0 else 0.0
        t["closing_cr"] = abs(cl_raw) if cl_raw < 0 else 0.0
        if "cls" not in t:
            l = ledgers.get(name, {})
            from parser import classify_ledger
            t["cls"] = classify_ledger(name, l.get("parent",""), groups)

    # ── 1. SUMMARY ──────────────────────────────────────────────────────────
    if "summary" in selected_sheets:
        ws = wb.create_sheet("Summary")
        ws.sheet_properties.tabColor = C_ACCENT
        total_income  = sum(t["closing_cr"]-t["closing_dr"] for t in trial_balance.values() if t["cls"] in PL_INCOME)
        total_expense = sum(t["closing_dr"]-t["closing_cr"] for t in trial_balance.values() if t["cls"] in PL_EXPENSE)
        net_profit    = total_income - total_expense
        total_assets  = sum(t["closing_dr"]-t["closing_cr"] for t in trial_balance.values() if t["cls"] in BS_ASSET)

        r = write_company_header(ws, company, "Audit Pack Summary Dashboard", f"{entity_type} · {period}")
        metrics = [
            ("Revenue from Operations", total_income),
            ("Total Expenses",          total_expense),
            ("Net Profit / (Loss)",     net_profit),
            ("Total Assets",            total_assets),
            ("Total Vouchers",          len(vouchers)),
            ("Total Ledgers",           len(ledgers)),
            ("Sheets in Pack",          total_sheets),
        ]
        write_col_headers(ws, r, ["Metric", "Value"], [40, 22])
        for i, (lbl, val) in enumerate(metrics, 1):
            is_tot = lbl in ("Net Profit / (Loss)", "Total Assets")
            write_data_row(ws, r+i, [lbl, val], is_total=is_tot)
        done += 1; log_fn("✔ Summary", pct())

    # ── 2. BALANCE SHEET ────────────────────────────────────────────────────
    if "balance_sheet" in selected_sheets:
        ws = wb.create_sheet("Balance_Sheet")
        ws.sheet_properties.tabColor = "3B82F6"
        r = write_company_header(ws, company,
            f"Balance Sheet as at {to_date}",
            f"Schedule III · {entity_type} · All amounts in {currency}")
        write_col_headers(ws, r, ["Particulars", "Note", f"As at {to_date}"], [46, 10, 20])
        r += 1

        def bs_section(title, cls_set, is_liability=True):
            nonlocal r
            write_data_row(ws, r, [f"  {title}", "", ""], is_section=True); r += 1
            total = 0.0
            items = [(n, t) for n, t in trial_balance.items() if t["cls"] in cls_set]
            for name, t in sorted(items, key=lambda x: x[0]):
                val = (t["closing_cr"] - t["closing_dr"]) if is_liability else (t["closing_dr"] - t["closing_cr"])
                if abs(val) > 0.005:
                    write_data_row(ws, r, [f"    {name}", "", val], is_alt=(r%2==0)); r += 1
                    total += val
            write_data_row(ws, r, [f"  TOTAL — {title}", "", total], is_total=True); r += 1
            return total

        write_data_row(ws, r, ["I.  EQUITY AND LIABILITIES", "", ""], is_section=True); r += 1
        liab = 0
        liab += bs_section("Share Capital",             {"equity"},       True)
        liab += bs_section("Reserves & Surplus",        {"reserve"},      True)
        liab += bs_section("Long-term Borrowings",      {"lt_loan"},      True)
        liab += bs_section("Trade Payables",            {"creditor"},     True)
        liab += bs_section("Provisions",                {"provision"},    True)
        liab += bs_section("Duties & Taxes (GST/TDS)",  {"tax_gst"},      True)
        liab += bs_section("Branch / Division Balances",{"branch"},       True)
        liab += bs_section("Other Current Liabilities", {"current_liab"}, True)
        write_data_row(ws, r, ["TOTAL EQUITY & LIABILITIES", "", liab], is_total=True); r += 2

        write_data_row(ws, r, ["II. ASSETS", "", ""], is_section=True); r += 1
        assets = 0
        assets += bs_section("Property, Plant & Equipment", {"fixed_asset"},  False)
        assets += bs_section("Investments",                  {"investment"},   False)
        assets += bs_section("Deposits",                     {"deposit"},      False)
        assets += bs_section("Trade Receivables",            {"debtor"},       False)
        assets += bs_section("Cash & Bank Balances",         {"bank"},         False)
        assets += bs_section("Loans & Advances",             {"advance"},      False)
        assets += bs_section("TDS / Tax Receivable",         {"tax_tds","tax_it"}, False)
        assets += bs_section("Other Current Assets",         {"current_asset","other"}, False)
        write_data_row(ws, r, ["TOTAL ASSETS", "", assets], is_total=True); r += 1
        write_data_row(ws, r, ["Difference (Assets − Liabilities)", "", round(assets-liab, 2)]); r += 1
        done += 1; log_fn("✔ Balance Sheet (Schedule III)", pct())

    # ── 3. PROFIT & LOSS ────────────────────────────────────────────────────
    if "profit_loss" in selected_sheets:
        ws = wb.create_sheet("Profit_Loss")
        ws.sheet_properties.tabColor = "22C55E"
        r = write_company_header(ws, company,
            f"Statement of Profit & Loss for year ended {to_date}",
            f"Schedule III · {entity_type} · {currency}")
        write_col_headers(ws, r, ["Particulars", "Note", f"Year ended {to_date}"], [46, 10, 20])
        r += 1

        write_data_row(ws, r, ["INCOME", "", ""], is_section=True); r += 1
        rev_total = 0.0
        for name, t in trial_balance.items():
            if t["cls"] == "income":
                v = t["closing_cr"] - t["closing_dr"]
                if abs(v) > 0.005:
                    write_data_row(ws, r, [f"  {name}", "", v], is_alt=(r%2==0)); r += 1
                    rev_total += v
        write_data_row(ws, r, ["I.  Revenue from Operations", "Note 1", rev_total], is_total=True); r += 1

        other_total = 0.0
        for name, t in trial_balance.items():
            if t["cls"] == "other_income":
                v = t["closing_cr"] - t["closing_dr"]
                if abs(v) > 0.005:
                    write_data_row(ws, r, [f"  {name}", "", v], is_alt=(r%2==0)); r += 1
                    other_total += v
        write_data_row(ws, r, ["II.  Other Income", "Note 2", other_total], is_total=True); r += 1
        total_income = rev_total + other_total
        write_data_row(ws, r, ["III. TOTAL REVENUE (I + II)", "", total_income], is_total=True); r += 2

        write_data_row(ws, r, ["EXPENSES", "", ""], is_section=True); r += 1
        purch_t, empl_t, indir_t = 0.0, 0.0, 0.0
        for name, t in trial_balance.items():
            v = t["closing_dr"] - t["closing_cr"]
            if abs(v) < 0.005: continue
            if t["cls"] == "purchase":
                write_data_row(ws, r, [f"  {name}", "", v], is_alt=(r%2==0)); r += 1; purch_t += v
            elif t["cls"] == "empl_expense":
                write_data_row(ws, r, [f"  {name}", "", v], is_alt=(r%2==0)); r += 1; empl_t += v
            elif t["cls"] == "indir_expense":
                write_data_row(ws, r, [f"  {name}", "", v], is_alt=(r%2==0)); r += 1; indir_t += v
        write_data_row(ws, r, ["Cost of Materials / Services", "Note 3", purch_t], is_total=True); r += 1
        write_data_row(ws, r, ["Employee Benefits Expense",    "Note 4", empl_t],  is_total=True); r += 1
        write_data_row(ws, r, ["Other Expenses",               "Note 5", indir_t], is_total=True); r += 1
        total_exp = purch_t + empl_t + indir_t
        write_data_row(ws, r, ["IV.  TOTAL EXPENSES", "", total_exp], is_total=True); r += 2
        net = total_income - total_exp
        write_data_row(ws, r, ["V.   PROFIT BEFORE TAX (III − IV)", "", net], is_total=True); r += 1
        write_data_row(ws, r, ["VI.  Tax Expense", "Note 6", 0]); r += 1
        write_data_row(ws, r, ["VII. PROFIT FOR THE PERIOD", "", net], is_total=True)
        done += 1; log_fn("✔ Profit & Loss (Schedule III)", pct())

    # ── 4. TRIAL BALANCE ────────────────────────────────────────────────────
    if "trial_balance" in selected_sheets:
        ws = wb.create_sheet("Trial_Balance")
        cols = ["Account", "Classification", "Parent Group",
                "Op Dr", "Op Cr", "Trans Dr", "Trans Cr", "Closing Dr", "Closing Cr"]
        widths = [36, 18, 26, 14, 14, 14, 14, 14, 14]
        r = write_company_header(ws, company, "Trial Balance — All Ledgers", period)
        write_col_headers(ws, r, cols, widths); r += 1
        for i, (name, t) in enumerate(sorted(trial_balance.items(), key=lambda x: (ledgers.get(x[0],{}).get("parent",""), x[0]))):
            l = ledgers.get(name, {})
            write_data_row(ws, r, [
                name, t.get("cls",""), l.get("parent",""),
                t.get("op_dr",0) or "", t.get("op_cr",0) or "",
                t.get("dr",0) or "", t.get("cr",0) or "",
                t.get("closing_dr",0) or "", t.get("closing_cr",0) or "",
            ], is_alt=(i%2==0)); r += 1
        done += 1; log_fn(f"✔ Trial Balance: {len(trial_balance)} ledgers", pct())

    # ── 5. DAYBOOK ──────────────────────────────────────────────────────────
    if "daybook" in selected_sheets:
        ws = wb.create_sheet("Daybook")
        cols = ["Date","Vch No","Vch Type","Account","Narration","Debit","Credit","Classification"]
        widths = [12, 16, 14, 30, 44, 15, 15, 16]
        r = write_company_header(ws, company, f"Daybook — {period}", f"{len(vouchers):,} voucher lines")
        write_col_headers(ws, r, cols, widths); r += 1
        for i, v in enumerate(vouchers):
            write_data_row(ws, r, [
                v.get("date",""), v.get("vch_no",""), v.get("vch_type",""),
                v.get("account",""), v.get("narration",""),
                v.get("debit",0) or "", v.get("credit",0) or "",
                v.get("cls",""),
            ], is_alt=(i%2==0)); r += 1
        done += 1; log_fn(f"✔ Daybook: {len(vouchers)} lines", pct())

    # ── 6. SALES REGISTER ───────────────────────────────────────────────────
    if "sales_reg" in selected_sheets:
        sales = [v for v in vouchers if v.get("cls") == "income" and v.get("credit",0) > 0]
        by_vch = {}
        for v in sales:
            k = f"{v['date']}|{v.get('vch_no','')}"
            if k not in by_vch:
                by_vch[k] = {"date":v["date"],"vch_no":v.get("vch_no",""),"vch_type":v.get("vch_type",""),
                             "account":v.get("account",""),"narration":v.get("narration",""),"total":0.0}
            by_vch[k]["total"] += v.get("credit",0)
        ws = wb.create_sheet("Sales_Register")
        cols = ["Date","Vch No","Vch Type","Sales Account","Taxable Value","GST%","CGST","SGST","IGST","Invoice Total","Narration"]
        widths = [12,16,12,30,15,8,12,12,12,15,40]
        r = write_company_header(ws, company, f"Sales Register — {len(by_vch)} vouchers", period)
        write_col_headers(ws, r, cols, widths); r += 1
        grand_total = 0.0
        for i, rec in enumerate(by_vch.values()):
            taxable = round(rec["total"]/1.18, 2)
            gst     = round(rec["total"] - taxable, 2)
            write_data_row(ws, r, [rec["date"],rec["vch_no"],rec["vch_type"],rec["account"],
                taxable,18,round(gst/2,2),round(gst/2,2),0,round(rec["total"],2),rec["narration"]], is_alt=(i%2==0))
            r += 1; grand_total += rec["total"]
        write_data_row(ws, r, ["GRAND TOTAL","","","","","","","","",round(grand_total,2),""], is_total=True)
        done += 1; log_fn(f"✔ Sales Register: {len(by_vch)} vouchers", pct())

    # ── 7. SALES SUMMARY ────────────────────────────────────────────────────
    if "sales_summary" in selected_sheets:
        ws = wb.create_sheet("Sales_Summary")
        acc_map = {}
        for v in vouchers:
            if v.get("cls") == "income" and v.get("credit",0) > 0:
                acc = v["account"]
                if acc not in acc_map: acc_map[acc] = {"count":0,"total":0.0}
                acc_map[acc]["count"] += 1
                acc_map[acc]["total"] += v["credit"]
        cols = ["Sales Account","Voucher Count","Total (Incl. GST)","Taxable Value","GST Amount"]
        widths = [36,14,18,18,14]
        r = write_company_header(ws, company, "Sales Summary by Account", period)
        write_col_headers(ws, r, cols, widths); r += 1
        for i,(acc, d) in enumerate(sorted(acc_map.items(), key=lambda x:-x[1]["total"])):
            taxable = round(d["total"]/1.18,2)
            write_data_row(ws, r,[acc,d["count"],round(d["total"],2),taxable,round(d["total"]-taxable,2)],is_alt=(i%2==0))
            r += 1
        done += 1; log_fn(f"✔ Sales Summary: {len(acc_map)} accounts", pct())

    # ── 8. PURCHASE REGISTER ────────────────────────────────────────────────
    if "purchase_reg" in selected_sheets:
        exp = [v for v in vouchers if v.get("cls") in ("purchase","empl_expense","indir_expense") and v.get("debit",0)>0]
        by_ledger = {}
        for v in exp:
            acc = v["account"]
            if acc not in by_ledger: by_ledger[acc] = []
            by_ledger[acc].append(v)
        ws = wb.create_sheet("Purchase_Register")
        cols = ["Ledger / Account","Schedule","Date","Vch No","Vch Type","Narration","Amount (Dr)"]
        widths = [36,18,12,16,14,44,16]
        r = write_company_header(ws, company, f"Purchase & Expense Register", period)
        write_col_headers(ws, r, cols, widths); r += 1
        for acc, vlist in sorted(by_ledger.items()):
            total = sum(v.get("debit",0) for v in vlist)
            write_data_row(ws, r, [acc, vlist[0].get("cls",""), "", "", "", f"{len(vlist)} vouchers", total], is_section=True); r += 1
            for i, v in enumerate(vlist):
                write_data_row(ws, r, [f"  {v.get('vch_no','')}","",v.get("date",""),v.get("vch_no",""),
                    v.get("vch_type",""),v.get("narration",""),v.get("debit",0)], is_alt=(i%2==0)); r += 1
        done += 1; log_fn(f"✔ Purchase Register: {len(exp)} lines", pct())

    # ── 9. CUSTOMER 360 ──────────────────────────────────────────────────────
    if "customer_360" in selected_sheets:
        cust_map = {}
        for v in vouchers:
            if v.get("cls") == "debtor":
                acc = v["account"]
                if acc not in cust_map: cust_map[acc] = {"dr":0.0,"cr":0.0}
                cust_map[acc]["dr"] += v.get("debit",0)
                cust_map[acc]["cr"] += v.get("credit",0)
        ws = wb.create_sheet("Customer_360")
        cols = ["Customer","GSTIN","Opening Bal","Sales (Dr)","Receipts (Cr)","Adjustments","Closing Bal","Status"]
        widths = [36,18,14,14,14,12,16,14]
        r = write_company_header(ws, company, "Customer 360°", "Opening + Sales − Receipts = Closing")
        write_col_headers(ws, r, cols, widths); r += 1
        for i,(name, mv) in enumerate(sorted(cust_map.items(), key=lambda x:-x[1]["dr"])):
            l = ledgers.get(name,{})
            op = l.get("opening_balance",0)
            cl = round(op + mv["dr"] - mv["cr"], 2)
            status = "Outstanding" if cl>0 else ("Credit Bal" if cl<0 else "Nil")
            write_data_row(ws,r,[name,l.get("gstin",""),op,mv["dr"],mv["cr"],0,cl,status],is_alt=(i%2==0)); r+=1
        done += 1; log_fn(f"✔ Customer 360°: {len(cust_map)} customers", pct())

    # ── 10. SUPPLIER 360 ─────────────────────────────────────────────────────
    if "supplier_360" in selected_sheets:
        sup_map = {}
        for v in vouchers:
            if v.get("cls") == "creditor":
                acc = v["account"]
                if acc not in sup_map: sup_map[acc] = {"dr":0.0,"cr":0.0}
                sup_map[acc]["dr"] += v.get("debit",0)
                sup_map[acc]["cr"] += v.get("credit",0)
        ws = wb.create_sheet("Supplier_360")
        cols = ["Supplier","GSTIN","Opening Bal","Purchases (Cr)","Payments (Dr)","Adjustments","Closing Bal","Status"]
        widths = [36,18,14,14,14,12,16,12]
        r = write_company_header(ws, company, "Supplier 360°", "Opening + Purchases − Payments = Closing")
        write_col_headers(ws, r, cols, widths); r += 1
        for i,(name, mv) in enumerate(sorted(sup_map.items(), key=lambda x:-x[1]["cr"])):
            l = ledgers.get(name,{})
            op = l.get("opening_balance",0)
            cl = round(op + mv["cr"] - mv["dr"], 2)
            status = "Payable" if cl>0 else ("Debit Bal" if cl<0 else "Nil")
            write_data_row(ws,r,[name,l.get("gstin",""),op,mv["cr"],mv["dr"],0,cl,status],is_alt=(i%2==0)); r+=1
        done += 1; log_fn(f"✔ Supplier 360°: {len(sup_map)} suppliers", pct())

    # ── 11. DEBTORS AGING ────────────────────────────────────────────────────
    if "debtors" in selected_sheets:
        from datetime import datetime, date
        try:
            today = datetime.strptime(to_date, "%Y-%m-%d").date() if to_date else date.today()
        except Exception:
            today = date.today()
        ws = wb.create_sheet("Debtors_Analysis")
        cols = ["Customer","Invoice Date","Vch No","Invoice Amt","0-30","31-60","61-90","91-180",">180"]
        widths = [34,12,16,14,12,12,12,14,12]
        r = write_company_header(ws, company, "Debtors Aging Analysis", f"As at {to_date}")
        write_col_headers(ws, r, cols, widths); r += 1
        for i,v in enumerate(v for v in vouchers if v.get("cls")=="debtor" and v.get("debit",0)>0):
            try:
                parts = str(v.get("date","")).replace("/","-").split("-")
                if len(parts)==3 and len(parts[2])==4:
                    inv_date = date(int(parts[2]),int(parts[1]),int(parts[0]))
                else:
                    inv_date = date(int(parts[0]),int(parts[1]),int(parts[2]))
                days = (today - inv_date).days
            except Exception:
                days = 0
            buckets = [0,0,0,0,0]
            bi = 0 if days<=30 else 1 if days<=60 else 2 if days<=90 else 3 if days<=180 else 4
            buckets[bi] = v.get("debit",0)
            write_data_row(ws, r, [v["account"],v.get("date",""),v.get("vch_no",""),v.get("debit",0)]+buckets, is_alt=(i%2==0)); r+=1
        done += 1; log_fn("✔ Debtors Aging", pct())

    # ── 12. GST SUMMARY ──────────────────────────────────────────────────────
    if "gst_summary" in selected_sheets:
        ws = wb.create_sheet("GST_Summary")
        cols = ["GST Ledger","Op Dr","Op Cr","Trans Dr","Trans Cr","Closing Dr","Closing Cr"]
        widths = [38,13,13,13,13,13,13]
        r = write_company_header(ws, company, "GST Summary — All Tax Ledgers", period)
        write_col_headers(ws, r, cols, widths); r += 1
        for i,(name,t) in enumerate((n,t) for n,t in trial_balance.items() if t["cls"]=="tax_gst"):
            write_data_row(ws,r,[name,t.get("op_dr",0),t.get("op_cr",0),t.get("dr",0),t.get("cr",0),t.get("closing_dr",0),t.get("closing_cr",0)],is_alt=(i%2==0)); r+=1
        done += 1; log_fn("✔ GST Summary", pct())

    # ── 13. GST 3-WAY ────────────────────────────────────────────────────────
    if "gst_3way" in selected_sheets:
        months = ["Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","Jan","Feb","Mar"]
        ws = wb.create_sheet("GST_3Way_Recon")
        cols = ["Month","Taxable Sales","CGST","SGST","IGST","Total GST","Invoice Value"]
        widths = [10,16,13,13,13,14,16]
        r = write_company_header(ws, company, "GST 3-Way Reconciliation — Month-wise", period)
        write_col_headers(ws, r, cols, widths); r += 1
        for i, m in enumerate(months):
            mvs = [v for v in vouchers if fiscal_month(v.get("date","")) == i]
            sales = sum(v.get("credit",0) for v in mvs if v.get("cls")=="income")
            cgst  = sum(v.get("credit",0) for v in mvs if "cgst" in v.get("account","").lower())
            sgst  = sum(v.get("credit",0) for v in mvs if "sgst" in v.get("account","").lower())
            igst  = sum(v.get("credit",0) for v in mvs if "igst" in v.get("account","").lower())
            taxable = round(sales/1.18,2)
            write_data_row(ws, r, [m,taxable,round(cgst,2),round(sgst,2),round(igst,2),round(cgst+sgst+igst,2),round(sales,2)], is_alt=(i%2==0)); r+=1
        done += 1; log_fn("✔ GST 3-Way Recon", pct())

    # ── 14. TDS REGISTER ─────────────────────────────────────────────────────
    if "tds" in selected_sheets:
        ws = wb.create_sheet("TDS_Register")
        cols = ["TDS Ledger","Opening","Deducted (Cr)","Deposited (Dr)","Closing Payable"]
        widths = [42,13,15,15,16]
        r = write_company_header(ws, company, "TDS Register — Section-wise", period)
        write_col_headers(ws, r, cols, widths); r += 1
        tds_items = [(n,t) for n,t in trial_balance.items() if "tds" in n.lower() or t.get("cls")=="tax_tds"]
        for i,(name,t) in enumerate(tds_items):
            op = t.get("op_cr",0)-t.get("op_dr",0)
            cl = t.get("closing_cr",0)-t.get("closing_dr",0)
            write_data_row(ws,r,[name,round(op,2),round(t.get("cr",0),2),round(t.get("dr",0),2),round(cl,2)],is_alt=(i%2==0)); r+=1
        done += 1; log_fn(f"✔ TDS Register: {len(tds_items)} ledgers", pct())

    # ── 15. EDIT LOG — VOUCHERS ──────────────────────────────────────────────
    if "editlog_vch" in selected_sheets:
        ws = wb.create_sheet("Edit_Log_Vouchers")
        ws.sheet_properties.tabColor = "22C55E"
        vl = edit_log.get("voucher_logs",[]) if edit_log else []
        cols = ["Audit Date","Audit Time","User","Action","Vch Date","Vch Type","Vch No / GUID","Amount","Cancelled?","Deleted?"]
        widths = [13,10,22,12,12,16,28,15,10,10]
        tot_del  = sum(1 for x in vl if x.get("is_deleted"))
        tot_canc = sum(1 for x in vl if x.get("is_cancelled"))
        r = write_company_header(ws, company,
            "Voucher Edit Log — MCA Audit Trail Compliance",
            f"{period} · Events: {len(vl)} · Deleted: {tot_del} · Cancelled: {tot_canc}")
        write_col_headers(ws, r, cols, widths); r += 1
        if vl:
            for i,l in enumerate(sorted(vl, key=lambda x:x.get("audit_date",""))):
                write_data_row(ws,r,[
                    l.get("audit_date","—"), l.get("audit_time","—"), l.get("user","Unknown"),
                    l.get("action","—"), l.get("vch_date","—"), l.get("vch_type","—"),
                    l.get("entity_name","—"), l.get("amount",0) or "",
                    "Yes" if l.get("is_cancelled") else "", "Yes" if l.get("is_deleted") else "",
                ], is_alt=(i%2==0)); r+=1
        else:
            ws.cell(r,1,"⚠ No Edit Log data. Export from: Gateway → Display More Reports → Audit & Compliance → Edit Log → Ctrl+E → XML")
        done += 1; log_fn(f"✔ Edit Log – Vouchers: {len(vl)} events", pct())

    # ── 16. EDIT LOG — MASTERS ───────────────────────────────────────────────
    if "editlog_mst" in selected_sheets:
        ws = wb.create_sheet("Edit_Log_Masters")
        ws.sheet_properties.tabColor = "22C55E"
        ml = edit_log.get("master_logs",[]) if edit_log else []
        cols = ["Audit Date","Audit Time","User","Action","Master Type","Master Name","Parent / Group","Deleted?"]
        widths = [13,10,22,12,14,32,28,10]
        r = write_company_header(ws, company, "Master Edit Log — Ledger / Group / Stock Item Changes",
            f"{period} · Events: {len(ml)}")
        write_col_headers(ws, r, cols, widths); r += 1
        if ml:
            for i,l in enumerate(sorted(ml, key=lambda x:x.get("audit_date",""))):
                write_data_row(ws,r,[
                    l.get("audit_date","—"),l.get("audit_time","—"),l.get("user","Unknown"),
                    l.get("action","—"),l.get("master_type","—"),l.get("entity_name","—"),
                    l.get("parent","—"),"Yes" if l.get("is_deleted") else "",
                ],is_alt=(i%2==0)); r+=1
        else:
            ws.cell(r,1,"⚠ No Master Edit Log data found.")
        done += 1; log_fn(f"✔ Edit Log – Masters: {len(ml)} events", pct())

    # ── 17. DELETED VOUCHERS ─────────────────────────────────────────────────
    if "editlog_del" in selected_sheets:
        ws = wb.create_sheet("Deleted_Vouchers")
        ws.sheet_properties.tabColor = "EF4444"
        vl = edit_log.get("voucher_logs",[]) if edit_log else []
        deleted = [l for l in vl if l.get("is_deleted") or l.get("is_cancelled")]
        cols = ["Audit Date","Audit Time","Deleted/Cancelled By","Action","Vch Date","Vch Type","Vch No / GUID","Amount"]
        widths = [13,10,24,14,12,16,28,15]
        r = write_company_header(ws, company,
            "Deleted & Cancelled Vouchers — MCA Audit Trail",
            f"⚠ {len(deleted)} entries removed from books · {period}")
        write_col_headers(ws, r, cols, widths); r += 1
        if deleted:
            for i,l in enumerate(deleted):
                write_data_row(ws,r,[
                    l.get("audit_date","—"),l.get("audit_time","—"),l.get("user","Unknown"),
                    l.get("action","—"),l.get("vch_date","—"),l.get("vch_type","—"),
                    l.get("entity_name","—"),l.get("amount",0) or "",
                ],is_alt=(i%2==0)); r+=1
        elif vl:
            write_data_row(ws, r, ["✅ No deleted or cancelled vouchers found."])
        else:
            ws.cell(r,1,"⚠ No Edit Log data loaded.")
        done += 1; log_fn(f"✔ Deleted Vouchers: {len(deleted)} entries", pct())

    # ── 18. AUDIT TRAIL SUMMARY ──────────────────────────────────────────────
    if "editlog_stats" in selected_sheets:
        ws = wb.create_sheet("Audit_Trail_Summary")
        ws.sheet_properties.tabColor = "22C55E"
        vl = edit_log.get("voucher_logs",[]) if edit_log else []
        ml = edit_log.get("master_logs",[])  if edit_log else []
        r = write_company_header(ws, company, "Audit Trail Summary — MCA Section 143", period)
        write_col_headers(ws, r, ["Metric","Value"], [36, 18]); r += 1
        write_data_row(ws, r, ["Total Voucher Events", len(vl)]); r += 1
        write_data_row(ws, r, ["Total Master Events",  len(ml)]); r += 1
        write_data_row(ws, r, ["Vouchers Deleted",     sum(1 for x in vl if x.get("is_deleted"))]); r += 1
        write_data_row(ws, r, ["Vouchers Cancelled",   sum(1 for x in vl if x.get("is_cancelled"))]); r += 1
        # User-wise
        user_map = {}
        for entry in vl + ml:
            u = entry.get("user","Unknown")
            if u not in user_map: user_map[u]={"created":0,"altered":0,"deleted":0,"cancelled":0,"masters":0}
            a = (entry.get("action","")).lower()
            if entry in ml: user_map[u]["masters"] += 1
            elif "creat" in a: user_map[u]["created"] += 1
            elif "delet" in a: user_map[u]["deleted"] += 1
            elif "cancel" in a: user_map[u]["cancelled"] += 1
            else: user_map[u]["altered"] += 1
        r += 1
        write_data_row(ws, r, ["USER-WISE ACTIVITY",""], is_section=True); r += 1
        write_col_headers(ws, r, ["User","Created","Altered","Deleted","Cancelled","Masters","Total"],
            [28,12,12,12,12,14,12]); r += 1
        for i,(u,c) in enumerate(user_map.items()):
            total = c["created"]+c["altered"]+c["deleted"]+c["cancelled"]
            write_data_row(ws,r,[u,c["created"],c["altered"],c["deleted"],c["cancelled"],c["masters"],total],is_alt=(i%2==0)); r+=1
        done += 1; log_fn("✔ Audit Trail Summary", pct())

    # ── 19. MONTHLY MIS ──────────────────────────────────────────────────────
    if "monthly_mis" in selected_sheets:
        months = ["Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","Jan","Feb","Mar"]
        ws = wb.create_sheet("Monthly_MIS")
        cols = ["Month","Income","Expenses","Gross Profit","Employee Cost","Other Exp","Net Profit","Margin %"]
        widths = [10,16,16,16,16,16,16,12]
        r = write_company_header(ws, company, f"Monthly MIS — FY {from_date[:4]}-{to_date[2:4]}", period)
        write_col_headers(ws, r, cols, widths); r += 1
        t_inc=t_exp=t_net=0.0
        for i,m in enumerate(months):
            mvs = [v for v in vouchers if fiscal_month(v.get("date","")) == i]
            inc  = sum(v.get("credit",0) for v in mvs if v.get("cls") in PL_INCOME)
            empl = sum(v.get("debit",0) for v in mvs if v.get("cls")=="empl_expense")
            oth  = sum(v.get("debit",0) for v in mvs if v.get("cls") in ("purchase","indir_expense"))
            net  = inc - empl - oth
            margin = f"{net/inc*100:.1f}%" if inc else "—"
            write_data_row(ws,r,[m,round(inc,2),round(empl+oth,2),round(inc-oth,2),round(empl,2),round(oth,2),round(net,2),margin],is_alt=(i%2==0)); r+=1
            t_inc+=inc; t_exp+=empl+oth; t_net+=net
        write_data_row(ws, r, ["TOTAL",round(t_inc,2),round(t_exp,2),"","","",round(t_net,2),""], is_total=True)
        done += 1; log_fn("✔ Monthly MIS: 12 months", pct())

    # ── 20. LEDGER-WISE ──────────────────────────────────────────────────────
    if "ledger_wise" in selected_sheets:
        by_ledger: dict = {}
        for v in vouchers:
            acc = v["account"]
            if acc not in by_ledger: by_ledger[acc] = []
            by_ledger[acc].append(v)
        ws = wb.create_sheet("Ledger_Wise")
        cols = ["Date","Vch No","Vch Type","Narration","Debit","Credit","Running Balance"]
        widths = [12,14,14,44,14,14,16]
        r = write_company_header(ws, company, "Ledger-wise Statement — Running Balances", period)
        write_col_headers(ws, r, cols, widths); r += 1
        for name, vlist in sorted(by_ledger.items()):
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            c = ws.cell(r, 1, f"── {name} ──")
            c.font = Font(name="Arial", size=9, bold=True, color=C_ACCENT)
            c.fill = fill(C_SEC_BG); r += 1
            bal = ledgers.get(name,{}).get("opening_balance",0.0)
            for i,v in enumerate(vlist):
                bal += v.get("debit",0) - v.get("credit",0)
                write_data_row(ws,r,[v.get("date",""),v.get("vch_no",""),v.get("vch_type",""),
                    v.get("narration",""),v.get("debit",0) or "",v.get("credit",0) or "",round(bal,2)],is_alt=(i%2==0)); r+=1
            write_data_row(ws, r, ["Closing Balance","","","","","",round(bal,2)], is_total=True); r += 1
        done += 1; log_fn(f"✔ Ledger-wise: {len(by_ledger)} ledgers", pct())

    # ── Freeze panes & auto-filter on all sheets ─────────────────────────────
    for ws in wb.worksheets:
        ws.freeze_panes = "A6"
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 90

    # ── Save to bytes ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
